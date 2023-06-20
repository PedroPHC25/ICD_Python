import openpyxl as xl
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
from openpyxl import styles as st

planilha = xl.Workbook()

planilha.create_sheet("Folha 1")

folha1 = planilha["Folha 1"]

folha1["A1"] = "Data"
folha1["B1"] = "Valor"
folha1["C1"] = "Descrição"

dados = ["10/10/2020", 900, "Aluguel"]
folha1.append(dados)

dados = [["10/10/2020", 900, "Aluguel"], ["10/10/2020", 900, "Aluguel"], ["10/10/2020", 900, "Aluguel"]]

for linha in dados:
    folha1.append(linha)

planilha.create_sheet("Folha 2")

folha2 = planilha["Folha 2"]

folha2.cell(row = 5, column = 5).value = "Data"
folha2.cell(row = 5, column = 6).value = "Valor"
folha2.cell(row = 5, column = 7).value = "Descrição"

comeco = (folha2.max_row + 1, folha2.min_column)

for linha, dado in enumerate(dados):
    for coluna, valor in enumerate(dado):
        folha2.cell(linha + comeco[0], coluna + comeco[1]).value = valor

tabela1 = Table(displayName = "Tabela1", ref = "A1:C5")
folha1.add_table(tabela1)

letra_1_coluna = get_column_letter(folha2.min_column)
n_1_linha = folha2.min_row
letra_ultima_coluna = get_column_letter(folha2.max_column)
n_ultima_linha = folha2.max_row

range_tabela = f"{letra_1_coluna}{n_1_linha}:{letra_ultima_coluna}{n_ultima_linha}"

tabela2 = Table(displayName = "Tabela2", ref = range_tabela)
folha2.add_table(tabela2)

fonte = st.Font(name = "Comic Sans", size = 15, bold = True, color = "FF0000")
folha1["A1"].font = fonte

preenchimento = st.PatternFill("solid", fgColor = "0000FF")
folha1["A1"].fill = preenchimento

folha1["E1"] = "=SUM(B2:B5)"

folha1.protection.sheet = True
folha1.protection.password = "123"

planilha.save("aula8.xlsx")