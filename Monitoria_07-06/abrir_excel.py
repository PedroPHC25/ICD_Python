import openpyxl as xl

planilha = xl.load_workbook("aula8.xlsx")

# print(planilha.sheetnames)

folha1 = planilha["Folha 1"]

dados = []

range_tabela = (folha1.min_row, folha1.min_column, folha1.max_row, folha1.max_column)

for linha in range(range_tabela[0], range_tabela[2] + 1):
    linha_atual = []
    for coluna in range(range_tabela[1], range_tabela[3] + 1):
        valor = folha1.cell(linha, coluna).value
        linha_atual.append(valor)
    dados.append(linha_atual)

print(dados)